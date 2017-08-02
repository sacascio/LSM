#!/usr/bin/env python

import getopt
import sys
import os
import openpyxl
import string
from openpyxl.workbook import workbook
from IPy import IP
import re
from collections import Counter
import xlsxwriter

# Prepared by Salvatore Cascio, Cisco Systems
# May 12, 2017
# Updated August 1, 2017

# Script assumptions:
#    1. Input file must be xls or xlsx, each field in pipe seperated format all in column A
#    2. Input data must be the first worksheet

def printDups(pstruct):
   fwb = xlsxwriter.Workbook('dup_pids.xlsx')
   formatwb = fwb.add_format()
   formatwb.set_center_across()
   
   formatwbheader = fwb.add_format()
   formatwbheader.set_bold()
   formatwbheader.set_center_across()
   formatwbheader.set_bg_color('yellow')

   fws = fwb.add_worksheet('nj2')
   fws.freeze_panes(1,0)
   fws.set_column(0,5,15)
   fws.center_horizontally()
   fws.write('A1','RFGW_NAME', formatwbheader)
   fws.write('B1','Frequency', formatwbheader)
   col = 'A'
   row = 2
   isValid = 1
   seen = 0
   result = []
   maxcount = -1
   
   for key in pstruct:
       for pid,count in Counter(pstruct[key]).items():
           if  int(pid) > 0 and count > 1:
               result.append(pid)

               if maxcount == -1:
                   maxcount = count

               else:
                   if count > maxcount:
                       maxcount = count

       if len(result) > 0:
           seen = 1
           data = key.split(",")
           rfgwname  = data[0]
           frequency = data[1]

           cell = col + str(row)
           fws.write(cell,rfgwname, formatwb)
           col = chr(ord(col) + 1)
           
           cell = col + str(row)
           fws.write_number(cell,int(frequency),formatwb)
           col = chr(ord(col) + 1)
          
           for i in result:
               cell = col + str(row)
               fws.write_number(cell,int(i),formatwb)
               col = chr(ord(col) + 1)
               
           isValid = 0
           result = []

           col = 'A'
           row = row + 1


   # add number of column headers - should equal max number of collided PIDs per carrier
   if maxcount < 2:
       if os.path.isfile('dup_pids.xlsx'):
           fwb.close()
           os.remove('dup_pids.xlsx')
   else:
        print "See dup_pids.xlsx - Duplicate PIDs per transport/RFGW-1 found"
        col = 'C'
        num = 1
        
        while num <= maxcount:
               cell   = col + str(1)
               pidnum = "PID " +  str(num)
               fws.write(cell,pidnum,formatwbheader)
               col = chr(ord(col) + 1)
               num = num +1
            
       
   fwb.close()
   return isValid

def addtouniqpids(pstruct,gwname,frequency,pk,nds,nagra,pmt,pcr,sessmac,mac):
    frequency = int(frequency) / 1000000
    mykey = gwname + "," + str(frequency)
    if sessmac.has_key(mac):
        #print "Key already exists, %s)" % ( mac )
        return 0

    sessmac[mac] = 1

    if  pstruct.has_key(mykey):
        currlist = pstruct[mykey]
        currlist = list(currlist)
        currlist.extend([pk,nds,nagra,pmt,pcr]) 
        pstruct[mykey] = (currlist)
    else:
        pstruct[mykey] = (pk,nds,nagra,pmt,pcr)

def isValidFreq(frequency):
	if ( (570000000 <= int(frequency) <= 963000000) and (int(frequency) % 3) ) == 0:
		return 1
	else:
		print('Invalid frequency provided, %s' % frequency)
		return 0


def isValidType(ptype):
    if bool(re.search('Audio|Video|Private', ptype)) == True:
        return 1
    else:
        print ('Invalid PID type provided.  Expected Audio, Video or Private.  Provided %s' % ptype)
        return 0
    

def isValidMod(modulation):
    if modulation == 'Qam256':
        return 1
    else:
        print('Invalid QAM modulation provided.  Expected Qam256, Provided %s' % modulation)
        return 0
    
    
def isNum(*args):
    isValid = 1
    all = args
    for a in all:
        isvnum = a.isnumeric()
        if isvnum == False:
            print (a + " is not a valid Number" )
            isValid = 0
        

    return isValid

def isValidIP(*args):
    isValid = 1
    all = args
    for a in all:
        try:
            IP(a)
        except:
            print (a + ' is not a valid IPv4 address')
            isValid = 0
            
    return isValid
     

def isValidSessId(macsessid):
    isValid = 1
    sessmac = macsessid.split("/")
    isValidnum = isNum(sessmac[1])
    if isValidnum == 0:
        print(sessmac[1] + " is not a valid Number, part of session MAC " + macsessid)
        
    if len(sessmac[0]) != 12:
            print(sessmac[0] + ' is not valid.  MAC address length is %s. Expecting 12' % len(sessmac[0]))
            isValid = 0
    else:
        if all(c in string.hexdigits for c in sessmac[0]) == False:
            print(sessmac[0] + " is not a valid MAC address, part of session mac " + macsessid)
            isValid = 0
        
    return isValid
    
  

def main(argv):
    writelog = 1
    pstruct = {}
    sessmac = {}

    try:
        opts,args = getopt.getopt(argv,"hf:l:",["file=","location="])
    except getopt.GetoptError as err:
        print str(err)
        sys.exit(2)
    else:
        for opt,arg in opts:
            if opt == '-h':
                print sys.argv[0] + " -f|--file <file_name>"
                sys.exit(1)
            elif opt in ( "-f", "--file"):
                filename = arg
            elif opt in ( "-l", "--location"):
                location = arg
            else:
                assert False, "Unknown"
                sys.exit(2)

    if len(argv) == 0:
        print "Usage: " +  sys.argv[0] + " -f|--file <file_name> -l|--location <location> No arguments given"
        sys.exit(1)

    try:
        filename
    except NameError:
        print "Filename not specified (-f or --file)"
        sys.exit(1)
    
    try:
        location
    except NameError:
        print "Location not specified (-l or --location)"
        sys.exit(1)
    
    if not os.path.isfile(filename):
        print('Filename ' + filename + " does not exist")
        sys.exit(4)
        
    workbook = openpyxl.load_workbook(filename)
    
    sessionsheet = workbook.worksheets[0]
    
    f = open(location + '.txt', 'w')
    
    for row in range (1, sessionsheet.max_row+1):
        #sessionid,macsessid,gwname,gwip,input_gbe,tsid,frequency,inudp,outudp,inmpeg,outmpeg,bw,ingbeip,outgbeip,pk,nds,nagra,sn,modulation,inpid,outpid,ptype,pmt,pcr = (sessionsheet['A' + str(row)].value).split("|")

        data = (sessionsheet['A' + str(row)].value).split("|")
        f.write(sessionsheet['A' + str(row)].value)
        if sessionsheet['A' + str(row)].value[-1] != "|":
            f.write("|\n")
        else:
            f.write("\n")

        validnums     = isNum(data[0],data[4],data[5],data[6],data[7],data[8],data[9],data[10],data[11],data[14],data[15],data[16],data[19],data[20],data[22],data[23])
        validsessid   = isValidSessId(data[1])
        validip       = isValidIP(data[3],data[12],data[13])
        validMod      = isValidMod(data[18])
        validtype     = isValidType(data[21])
        validFreq     = isValidFreq(data[6]) 
    
        #pattern = re.search("visible",data[17],re.IGNORECASE)
        vwcount = len(re.findall("visible|vw", data[17],re.IGNORECASE))

        if data[21] == 'Video' and vwcount == 0:
            #addtouniqpids(pstruct,data[2],data[6],data[10],data[14],data[15],data[16],data[22],data[23],sessmac,data[1])
            addtouniqpids(pstruct,data[2],data[6],data[14],data[15],data[16],data[22],data[23],sessmac,data[1])

        
        if ( (not validnums) | (not validsessid) | (not validip) | (not validMod) | (not validtype) | (not validFreq ) ):
            print('Row %s has failures' % row)
            print('************************************************************************\n\n')
            writelog = 0
            
    f.close()
    workbook.close()
    
    isValidStruct = printDups(pstruct)
    
    if writelog == 0 or isValidStruct == 0:
        os.remove(location + '.txt')
        print('input file for LSM NOT created!!!')
    else:
        print('input file for LSM created: %s.txt' ) % ( location )
    
        
if __name__ == '__main__':
    main(sys.argv[1:])
